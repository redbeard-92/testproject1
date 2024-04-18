import openpyxl
import subprocess
import socket


def check_ping(ip_address):
    try:
        result = subprocess.run(['ping', '-n', '1', ip_address], capture_output=True, text=True, timeout=5)
        return result.returncode == 0
    except Exception as e:
        print(f"Ошибка при выполнении ping для {ip_address}: {e}")
        return False


def get_hostname(ip_address):
    try:
        result = subprocess.check_output(['ping', '-a', ip_address], stderr=subprocess.STDOUT, universal_newlines=True)
        hostname = result.split(' ')[1].strip('[]')
        return hostname
    except subprocess.CalledProcessError:
        return None


def detect_os(ip_address):
    try:
        result = subprocess.run(['ping', '-n', '1', ip_address], capture_output=True, text=True, timeout=5)
        if result.returncode == 0:
            if "Reply" in result.stdout:
                return "Windows"
            else:
                return "Linux"
    except Exception as e:
        print(f"Ошибка при определении операционной системы для IP {ip_address}: {e}")
    return "Unknown"


def update_os_from_hosts_as(hosts_file, hosts_as_file):
    # Открываем файлы
    wb_hosts = openpyxl.load_workbook(hosts_file)
    wb_hosts_as = openpyxl.load_workbook(hosts_as_file)

    # Получаем активные листы
    sheet_hosts = wb_hosts.active
    sheet_hosts_as = wb_hosts_as.active

    # Создаем словарь для быстрого доступа к операционным системам по хостнеймам
    os_dict = {}
    for row in sheet_hosts_as.iter_rows(min_row=2, values_only=True):
        os_dict[row[0]] = row[4]

    # Сравниваем и обновляем операционные системы
    for row in range(2, sheet_hosts.max_row + 1):
        hostname = sheet_hosts.cell(row=row, column=3).value
        os = os_dict.get(hostname)
        if os:
            sheet_hosts.cell(row=row, column=5).value = os

    # Сохраняем изменения
    wb_hosts.save(hosts_file)


def process_devices(sheet):
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        ip_address = sheet.cell(row=row, column=2).value
        if ip_address:
            print(f"Пингуется IP: {ip_address}")
            try:
                if check_ping(ip_address):
                    sheet.cell(row=row, column=4).value = "Доступен"
                else:
                    sheet.cell(row=row, column=4).value = "Недоступен"

                hostname = get_hostname(ip_address)
                if hostname:
                    sheet.cell(row=row, column=3).value = hostname.encode('ascii', 'ignore').decode('ascii')
                else:
                    sheet.cell(row=row, column=3).value = "Не удалось получить хостнейм"

                os = detect_os(ip_address)
                sheet.cell(row=row, column=5).value = os
            except Exception as e:
                print(f"Ошибка при обработке устройства с IP {ip_address}: {e}")

hosts_file = 'hosts.xlsx'
wb_hosts = openpyxl.load_workbook(hosts_file)
sheet_hosts = wb_hosts.active
process_devices(sheet_hosts)
update_os_from_hosts_as(hosts_file, 'hostsAs.xlsx')

print("Готово!")
